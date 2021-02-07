import openpyxl
ticker_and_company = dict()
ticker_and_ltp = dict()

workbook = openpyxl.load_workbook("./data/STONKS V2.5Beta.xlsx")
sheet = workbook["STONKS"]
sheet = workbook.active
count = 0
for j in range(1, sheet.max_row):
	a = sheet["A" + str(j)]
	b = sheet["B" + str(j)]
	if b.value == "#N/A": continue
	ticker_and_company[b.value] = a.value

for j in range(1, sheet.max_row):
	b = sheet["B" + str(j)]
	c = sheet["C" + str(j)]
	if c.value == "#N/A": continue
	ticker_and_ltp[b.value] = c.value


